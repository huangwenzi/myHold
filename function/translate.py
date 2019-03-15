# 系统库
import os
from PIL import Image
from PySide2 import QtWidgets
from PySide2 import QtCore
from openpyxl import Workbook  # 新建时导入这个
from openpyxl import load_workbook  # 读取时导入这个
import re
import json
import pytoml as toml
import _thread
# 项目库
from tools.log_tool import log_tool
from tools.config_tool import config_tool
from configs.enums import enums

# 翻译的类
class Translate(QtWidgets.QWidget):
    # 翻译数据
    translate_data = {}
    # 加载翻译的锁
    load_flag = False
    
    # parent : 父窗口
    def __init__(self, parent):
        super(Translate, self).__init__(parent)
        # 设置模态窗口
        self.setWindowFlags(QtCore.Qt.Dialog)
        # self.setWindowModality(QtCore.Qt.WindowModal)

        # 界面设置
        self.cfg = config_tool.cfg_map["translate_windows"]
        cfg = self.cfg
        self.setWindowTitle(cfg["windows_name"])
        self.resize(cfg["windows_width"], cfg["windows_height"])

        # 界面部件
        # 标签
        # 目标目录
        self.lable_excel = QtWidgets.QLabel(self)
        self.lable_excel.setText("excel文件")
        self.lable_toml = QtWidgets.QLabel(self)
        self.lable_toml.setText("toml文件")
        self.lable_find = QtWidgets.QLabel(self)
        self.lable_find.setText("要查询的句子")

        # 文本输入
        # excel待转化文件
        self.lineEdit_excel = QtWidgets.QLineEdit(self)
        self.lineEdit_excel.resize(150, 30)
        self.lineEdit_excel.setText("1.xlsx")
        # 翻译toml文件
        self.lineEdit_toml = QtWidgets.QLineEdit(self)
        self.lineEdit_toml.resize(150, 30)
        # 要查询的句子
        self.lineEdit_find = QtWidgets.QLineEdit(self)
        self.lineEdit_find.resize(150, 30)
        # 显示的翻译
        self.textEdit_find = QtWidgets.QTextEdit(self)
        self.textEdit_find.resize(300, 150)

        # 按钮
        # excel转化toml
        self.button_excel_to_toml = QtWidgets.QPushButton(self)
        self.button_excel_to_toml.setText("加载excel")
        # 加载toml
        self.button_load_toml = QtWidgets.QPushButton(self)
        self.button_load_toml.setText("加载toml")
        # 查询按钮
        self.button_find = QtWidgets.QPushButton(self)
        self.button_find.setText("查询翻译")


        # 布局管理器
        self.layout = QtWidgets.QGridLayout(self)
        self.layout.addWidget(self.lable_excel, 1, 0, 1, 1)
        self.layout.addWidget(self.lineEdit_excel, 1, 1, 1, 1)
        self.layout.addWidget(self.button_excel_to_toml, 1, 2, 1, 1)
        self.layout.addWidget(self.lable_toml, 2, 0, 1, 1)
        self.layout.addWidget(self.lineEdit_toml, 2, 1, 1, 1)
        self.layout.addWidget(self.button_load_toml, 2, 2, 1, 1)
        self.layout.addWidget(self.lable_find, 4, 0, 1, 1)
        self.layout.addWidget(self.lineEdit_find, 4, 1, 1, 2)
        self.layout.addWidget(self.button_find, 4, 2, 1, 1)
        self.layout.addWidget(self.textEdit_find, 5, 0, 5, 3)

        self.setLayout(self.layout)

        # 设置连接信号
        self.button_excel_to_toml.clicked.connect(self.click_excel_to_toml)
        self.button_find.clicked.connect(self.click_find)

        
    # 点击excel转化toml按钮(这部分可以用线程去运行，不妨碍其他功能使用)
    def click_excel_to_toml(self):
        # 检查锁
        if self.load_flag:
            log_tool.log("quickKey, click_find, 正在加载翻译")
            return
        try:
            _thread.start_new_thread(self.load_excel)
        except:
            log_tool.log("quickKey, click_excel_to_toml, Error: 无法启动线程, load_excel")

        

    # 点击查找翻译按钮
    def click_find(self):
        # 检查锁
        if self.load_flag:
            log_tool.log("quickKey, click_find, 正在加载翻译")
            return

        # 先清除显示
        self.textEdit_find.clear()
        # 获取文本
        find_str = self.lineEdit_find.text()

        # 是否找到的flag
        find_flag = False
        # 遍历查找包含的
        for tmp_key in self.translate_data:
            if tmp_key.find(find_str) != -1:
                find_flag = True
                tmp_data = self.translate_data[tmp_key]
                # 暂时只要韩文翻译
                self.textEdit_find.insertPlainText(tmp_data["翻译内容"] + "\n")
                self.textEdit_find.insertPlainText(tmp_data["韩文翻译"] + "\n")

        if not find_flag:
            self.textEdit_find.setPlainText("翻译不存在")

    # 加载excel的翻译
    def load_excel(self):
        # 加锁
        self.load_flag = True

        # 获取文本
        excel_name = self.lineEdit_excel.text()

        # 检查目标目录的输入
        if os.path.exists(excel_name) == False:
            msg = "目标目录:" + excel_name + " 不存在"
            log_tool.log(msg, enums.log_type.both, self)
            return

        # 读取excel数据
        str_data = {}
        workbook = load_workbook(excel_name)
        sheet = workbook["翻译"]
        # 获取对应的翻译语种
        col_num = sheet.max_column
        language_arr = []   # 语种的集合
        for index in range(1,col_num + 1):
            language_arr.append(sheet.cell(row=1, column=index).value)
        # 读取每一句对应的翻译
        row_num = sheet.max_row
        for index in range(2,row_num + 1):
            # 读取没一行的对应语种
            tmp_data = {}
            for index_col in range(1,col_num + 1):
                tmp_data[language_arr[index_col-1]] = sheet.cell(row=index, column=index_col).value

            str_data[tmp_data[language_arr[0]]] = tmp_data
        
        self.translate_data = str_data

        log_tool.log("加载excel翻译完毕", enums.log_type.both, self)

        # 解除锁
        self.load_flag = False